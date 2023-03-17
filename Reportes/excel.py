#----------------------------------------------------------------------------------------------------------------#
# Importaciones necesarias
#----------------------------------------------------------------------------------------------------------------#

from .functions import *

from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

#----------------------------------------------------------------------------------------------------------------#
# Archivos Excel descargables
#----------------------------------------------------------------------------------------------------------------#

class ReportesExcel():

    # ------------------------------ #
    # Ajustes del título y cabeceras #
    # ------------------------------ #

    def ajustes_excel(ws, valorB2, lista_letras, bordeado):
        lista_cabeceras = ['ID','Cliente','Obra','Fecha','Operador','Hora Ingreso','Hora Término','Horas Arriendo','Horómetro Inicial','Horómetro Final','Horómetro Total','Equipo N°','Hora Mínima','Accesorios','Observación']
        centrado        = Alignment(horizontal = "center", vertical = "center")

        # Dimensiones generales
            # filas
        ws.row_dimensions[1].height = 10
        ws.row_dimensions[2].height = 30
        ws.row_dimensions[4].height = 15
            # columnas
        ws.column_dimensions['A'].width = 2
        ws.column_dimensions['B'].width = 5
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 11
        ws.column_dimensions['F'].width = 35
        ws.column_dimensions['G'].width = 13
        ws.column_dimensions['H'].width = 13
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 16
        ws.column_dimensions['K'].width = 16
        ws.column_dimensions['L'].width = 16
        ws.column_dimensions['M'].width = 10
        ws.column_dimensions['N'].width = 12
        ws.column_dimensions['O'].width = 30
        ws.column_dimensions['P'].width = 50

        # Título y cabeceras
        ws.merge_cells('B2:P2')
        ws.merge_cells('B3:P3')
        
        b2 = ws['B2']

        b2.value     = valorB2
        b2.font      = Font(bold=True, size=24)
        b2.alignment = centrado
        b2.fill      = PatternFill(start_color='FFF1D4', end_color='FFF1D4', fill_type='solid')

        for letra in lista_letras:
            ws[letra + '2'].border = bordeado

            ws[letra + '4'].value       = lista_cabeceras[lista_letras.index(letra)]
            ws[letra + '4'].fill        = PatternFill(start_color='7299FA', end_color='7299FA', fill_type='solid')
            ws[letra + '4'].font        = Font(color='FFFFFF', bold=True)
            ws[letra + '4'].border      = bordeado
            ws[letra + '4'].alignment   = centrado

    # ------------------------------ #
    # Ajustes de los reportes        #
    # ------------------------------ #
    def ajustes_reporte(ws, lista_letras, bordeado, reportes, contador):
        # Reportes
        x = True
        for r in reportes:
            detalle    = AccesorioReporte.objects.filter(r_id_reporte = r.id_reporte).values_list()
            accesorios = ''
            cont       = 1
            for d in detalle:
                accesorio = Accesorio.objects.get(id_accesorio = d[1])
                accesorios += str(accesorio.accesorio_minicargador)
                if cont < len(detalle):
                    accesorios += ', '
                    cont += 1
                else:
                    accesorios += '.'

            ws.row_dimensions[contador].height = 25
            p = r.u_p_id_persona.p_id_persona
            lista_reporte = [str(r.id_reporte),str(r.cliente),str(r.obra),str(r.fecha),str(p.nombres) + " " + str(p.apellido_paterno) + " " + str(p.apellido_materno),str(r.hora_ingreso),str(r.hora_termino),str(r.horas_arriendo),str(r.horometro_inicial),str(r.horometro_final),str(r.horometro_total),str(r.equipo_numero),str(r.hora_minima),accesorios,str(r.observaciones)]

            for letra in lista_letras:
                ws[letra + str(contador)].border = bordeado
                ws[letra + str(contador)].value  = lista_reporte[lista_letras.index(letra)]

            if x == True:
                x = False
            else:
                for letra in lista_letras:
                    ws[letra + str(contador)].fill = PatternFill(start_color='DBDBDB', end_color='DBDBDB', fill_type='solid')
                x = True

            contador +=1

    # ------------------------------ #
    # Base de los archivos excel     #
    # ------------------------------ #
    def base_excel(variasHojas,   # Booleano para determinar si son varias hojas.
                   lista_hojas,   # Listado de las hojas si hay varias de estas.
                   titulo,        # Título de la hoja en caso de sólo tener una.
                   valorB2,       # Valor de la celda B2 del excel.
                   Reportes,      # Lista de los reportes necesarios.
                   nombre_archivo # Nombre del archivo excel.
        ):
        wb = Workbook()

        lista_letras = 'BCDEFGHIJKLMNOP'
        bordeado     = Border(top = Side(border_style = "double"),right = Side(border_style = "double"),bottom = Side(border_style = "double"),left = Side(border_style = "double"))
        contador     = 5

        varias_hojas = variasHojas

        if varias_hojas:
            # hoja_activa = True
            # for hoja in lista_hojas:
            #     contador = 5
            #     # reportes = Reportes
            pass
        else:
            ws       = wb.active
            ws.title = titulo
            ReportesExcel.ajustes_excel(ws, valorB2, lista_letras, bordeado)
            ReportesExcel.ajustes_reporte(ws, lista_letras, bordeado, Reportes, contador)

        response = HttpResponse(content_type = "application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo + '.xlsx')
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
    
# ----------------------------------------------------------------------------------------------------------------------------------------
# Archivos Excel
# ----------------------------------------------------------------------------------------------------------------------------------------
    # ------------------------------ #
    # Reportes Totales               #
    # ------------------------------ #
    def reportes(self):
        varias_hojas   = False
        lista_hojas    = ''
        titulo         = 'Reportes'
        valor_b2       = 'Reportes'
        reportes       = Reporte.objects.filter(valido = 1).order_by('id_reporte')
        nombre_archivo = 'Reportes_Totales'
        return ReportesExcel.base_excel(varias_hojas, lista_hojas, titulo, valor_b2, reportes, nombre_archivo)

    # ------------------------------ #
    # Reportes de un operador        #
    # ------------------------------ #
    def reporte_operador(self, pk):
        usuario         = Usuario.objects.get(p_id_persona = pk)
        p               = usuario.p_id_persona
        nombreCompleto  = str(p.nombres) + " " + str(p.apellido_paterno) + " " + str(p.apellido_materno)
        nombre_completo = str(p.nombres) + "_" + str(p.apellido_paterno) + "_" + str(p.apellido_materno)

        varias_hojas   = False
        lista_hojas    = ''
        titulo         = nombreCompleto
        valor_b2       = 'Reportes de ' + nombreCompleto
        reportes       = Reporte.objects.filter(u_p_id_persona = usuario, valido = 1).order_by('-fecha')
        nombre_archivo = 'Reportes_'+ nombre_completo
        return ReportesExcel.base_excel(varias_hojas, lista_hojas, titulo, valor_b2, reportes, nombre_archivo)
# ----------------------------------------------------------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------------------------------

    def reportes_mes(self):
        wb          = Workbook()
        lista_meses = ['Enero','Febrero','Marzo','Abril','Mayo','Junio','Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']

        t = True
        lista_letras = 'BCDEFGHIJKLMNOP'
        bordeado = Border(top = Side(border_style = "double"),right = Side(border_style = "double"),bottom = Side(border_style = "double"),left = Side(border_style = "double"))

        for mes in lista_meses:
            contador = 5
            reportes = operacionesFechas.reporte_mes(lista_meses.index(mes)+1)
            if t == True:
                t = False
                ws = wb.active
                ws.title = mes
            else:
                ws = wb.create_sheet(mes)

            valor_B2 = 'Reportes de ' + mes
            ReportesExcel.ajustes_excel(ws, valor_B2, lista_letras, bordeado)
            ReportesExcel.ajustes_reporte(ws, lista_letras, bordeado, reportes, contador)
                

        nombre_archivo = 'Reportes_Mes.xlsx'

        response = HttpResponse(content_type = "application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response



    def reportes_persona(self):
        wb          = Workbook()
        operadores  = Usuario.objects.filter(r_id_rol = 1)

        t = True
        lista_letras = 'BCDEFGHIJKLMNOP'
        bordeado = Border(top = Side(border_style = "double"),right = Side(border_style = "double"),bottom = Side(border_style = "double"),left = Side(border_style = "double"))

        # Crea una hoja para cada operador y llena con los respectivos datos.
        for o in operadores:
            persona = o.p_id_persona
            reportes = Reporte.objects.filter(u_p_id_persona = o, valido = 1).order_by('fecha')
            nombre_completo = persona.nombres + ' ' + persona.apellido_paterno + ' ' + persona.apellido_materno
            contador = 5
            # Trabaja con la primera hora y le pone nombre, en el siguiente recorrido crea una nueva hoja con otro nombre
            if t:
                t = False
                ws = wb.active
                ws.title = nombre_completo
            else:
                ws = wb.create_sheet(nombre_completo)

            
            valor_B2 = 'Reportes de ' + nombre_completo
            ReportesExcel.ajustes_excel(ws, valor_B2, lista_letras, bordeado)
            ReportesExcel.ajustes_reporte(ws, lista_letras, bordeado, reportes, contador)
                

        nombre_archivo = 'Reportes_Personas.xlsx'

        response = HttpResponse(content_type = "application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response